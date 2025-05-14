import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border


class ExcelService:
    @staticmethod
    def convert_excel_to_json(excel_file):
        try:
            df = pd.read_excel(excel_file)
            order_data = []
            for index, row in df.iterrows():
                row_dict = row.to_dict()
                if row_dict['amount'] != 0:
                    order_data.append(row_dict)
            return order_data
        except Exception as e:
            print(e)
            response_object = {
                'status': 'fail',
                'message': 'Try again'
            }
            return response_object, 500

    @staticmethod
    def make_summary_excel(summary):
        # Создаем DataFrame из данных
        df = pd.DataFrame(summary)
        price_columns = [col for col in df.columns if
                         col not in ['name', 'amount'] and not re.match(r'comment_\d+', col)]

        # Преобразуем цены в float
        for col in price_columns:
            df[col] = df[col].astype(float)

        # Создаем Excel-файл
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Summary')
            workbook = writer.book
            worksheet = writer.sheets['Summary']

            # Определяем минимальные цены для каждой строки (игнорируя None и 0)
            min_prices = []
            for idx, row in df.iterrows():
                prices = [row[col] for col in price_columns if not pd.isna(row[col]) and row[col] != 0]
                print(prices)
                min_price = min(prices) if prices else None
                min_prices.append(min_price)
                print(min_price, min_prices)


            # Определяем индексы колонок цен в Excel (учитывая, что первая строка - заголовки)
            col_indices = {col: df.columns.get_loc(col) + 1 for col in price_columns}

            # Мягкий зеленый цвет для выделения
            light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

            # Применяем условное форматирование для каждой строки
            for row_idx, min_price in enumerate(min_prices, start=2):  # Начинаем с 2-й строки (1 - заголовок)
                if min_price is None:
                    continue  # Пропускаем строки без валидных цен
                for col_name in price_columns:
                    col_idx = col_indices[col_name]
                    cell_value = df.at[row_idx - 2, col_name]  # row_idx - 2 для индекса в DataFrame
                    if cell_value == min_price and cell_value != 0:
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = light_green_fill

        output.seek(0)
        return output

    @staticmethod
    def from_summary(file):
        df = pd.read_excel(file)
        order_data = []
        for index, row in df.iterrows():
            row_dict = row.to_dict()
            if row_dict['amount'] != 0:
                order_data.append(row_dict)
        summary = {item['name']: {key: value for key, value in item.items() if
                                  not re.match(r"^comment_\d+$", key)
                                  and key not in ['name', 'amount'] and not pd.isna(value)} for item in
                   order_data}
        return summary

    @staticmethod
    def make_person_order_excel(summary):
        df = pd.DataFrame(summary)
        df['price'] = df['price'].astype(float)


        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Order')
        output.seek(0)
        return output
